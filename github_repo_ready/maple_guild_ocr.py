import os
import sys
import re
import json
import math
import tempfile
from io import BytesIO
from difflib import SequenceMatcher
from urllib.parse import urljoin, urlparse, parse_qs, urlencode, urlunparse

os.environ["PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK"] = "True"

import cv2
import numpy as np
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image as XLImage
from paddleocr import PaddleOCR

from PySide6.QtCore import QThread, Signal, Qt
from PySide6.QtWidgets import (
    QApplication,
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QPushButton,
    QComboBox,
    QLineEdit,
    QTextEdit,
    QMessageBox,
    QFileDialog,
    QListWidget,
    QSpinBox,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
)

BASE = "https://maplestory.nexon.com"

WORLD_BASE_URLS = {
    "에오스": "https://maplestory.nexon.com/N23Ranking/World/Guild?w=2&t=0",
    "헬리오스": "https://maplestory.nexon.com/N23Ranking/World/Guild?w=1&t=0",
    "핼리오스": "https://maplestory.nexon.com/N23Ranking/World/Guild?w=1&t=0",
    "오로라": "https://maplestory.nexon.com/N23Ranking/World/Guild?w=3&t=0",
    "레드": "https://maplestory.nexon.com/N23Ranking/World/Guild?w=4&t=0",
    "이노시스": "https://maplestory.nexon.com/N23Ranking/World/Guild?w=5&t=0",
    "유니온": "https://maplestory.nexon.com/N23Ranking/World/Guild?w=6&t=0",
    "스카니아": "https://maplestory.nexon.com/N23Ranking/World/Guild?w=7&t=0",
    "루나": "https://maplestory.nexon.com/N23Ranking/World/Guild?w=8&t=0",
    "제니스": "https://maplestory.nexon.com/N23Ranking/World/Guild?w=9&t=0",
    "크로아": "https://maplestory.nexon.com/N23Ranking/World/Guild?w=10&t=0",
    "베라": "https://maplestory.nexon.com/N23Ranking/World/Guild?w=11&t=0",
    "엘리시움": "https://maplestory.nexon.com/N23Ranking/World/Guild?w=12&t=0",
    "아케인": "https://maplestory.nexon.com/N23Ranking/World/Guild?w=13&t=0",
    "노바": "https://maplestory.nexon.com/N23Ranking/World/Guild?w=14&t=0",
    "챌린저스": "https://maplestory.nexon.com/N23Ranking/World/Guild?w=19&t=0",
    "챌린저스2": "https://maplestory.nexon.com/N23Ranking/World/Guild?w=20&t=0",
    "챌린저스3": "https://maplestory.nexon.com/N23Ranking/World/Guild?w=21&t=0",
    "챌린저스4": "https://maplestory.nexon.com/N23Ranking/World/Guild?w=22&t=0",
}

ROLE_WORDS = {"마스터", "부마스터", "길드원", "수습길드원", "가입대기"}

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/135.0.0.0 Safari/537.36"
    )
}

LIGHT_PURPLE_FILL = PatternFill(fill_type="solid", fgColor="E6D9F2")
LIGHT_GRAY_FILL = PatternFill(fill_type="solid", fgColor="EDEDED")

KOREAN_OCR_MODEL_NAME = "korean_PP-OCRv5_mobile_rec"
G_FLAG_VALUES = ("0", "1,000")

HIGH_CONFIDENCE_LOCK_THRESHOLD = 0.75
DUPLICATE_PENALTY_PER_EXTRA_USE = 0.08

CHOSUNG = ["ㄱ", "ㄲ", "ㄴ", "ㄷ", "ㄸ", "ㄹ", "ㅁ", "ㅂ", "ㅃ", "ㅅ", "ㅆ", "ㅇ", "ㅈ", "ㅉ", "ㅊ", "ㅋ", "ㅌ", "ㅍ", "ㅎ"]
JUNGSUNG = ["ㅏ", "ㅐ", "ㅑ", "ㅒ", "ㅓ", "ㅔ", "ㅕ", "ㅖ", "ㅗ", "ㅘ", "ㅙ", "ㅚ", "ㅛ", "ㅜ", "ㅝ", "ㅞ", "ㅟ", "ㅠ", "ㅡ", "ㅢ", "ㅣ"]
JONGSUNG = ["", "ㄱ", "ㄲ", "ㄳ", "ㄴ", "ㄵ", "ㄶ", "ㄷ", "ㄹ", "ㄺ", "ㄻ", "ㄼ", "ㄽ", "ㄾ", "ㄿ", "ㅀ", "ㅁ", "ㅂ", "ㅄ", "ㅅ", "ㅆ", "ㅇ", "ㅈ", "ㅊ", "ㅋ", "ㅌ", "ㅍ", "ㅎ"]


# -----------------------------
# 공통 유틸
# -----------------------------
def clean_text(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def normalize_key(text: str) -> str:
    if not text:
        return ""
    text = text.strip()
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[^가-힣A-Za-z0-9]", "", text)
    return text.lower()


def set_query_params(url: str, **kwargs) -> str:
    parsed = urlparse(url)
    query = parse_qs(parsed.query)

    for key, value in kwargs.items():
        if value is None:
            continue
        query[key] = [str(value)]

    new_query = urlencode(query, doseq=True)
    return urlunparse(
        (parsed.scheme, parsed.netloc, parsed.path, parsed.params, new_query, parsed.fragment)
    )


def decompose_hangul(text: str) -> str:
    result = []

    for ch in text:
        code = ord(ch)
        if 0xAC00 <= code <= 0xD7A3:
            base = code - 0xAC00
            cho = base // 588
            jung = (base % 588) // 28
            jong = base % 28

            result.append(CHOSUNG[cho])
            result.append(JUNGSUNG[jung])
            if JONGSUNG[jong]:
                result.append(JONGSUNG[jong])
        else:
            result.append(ch.lower())

    return "".join(result)


def mixed_name_similarity(a: str, b: str) -> float:
    na = normalize_key(a)
    nb = normalize_key(b)

    if not na or not nb:
        return 0.0

    raw_ratio = SequenceMatcher(None, na, nb).ratio()

    da = decompose_hangul(na)
    db = decompose_hangul(nb)
    jamo_ratio = SequenceMatcher(None, da, db).ratio()

    alnum_a = re.sub(r"[^A-Za-z0-9]", "", a).lower()
    alnum_b = re.sub(r"[^A-Za-z0-9]", "", b).lower()

    if alnum_a and alnum_b:
        alnum_ratio = SequenceMatcher(None, alnum_a, alnum_b).ratio()
    else:
        alnum_ratio = 0.0

    score = (0.35 * raw_ratio) + (0.55 * jamo_ratio) + (0.10 * alnum_ratio)
    return min(1.0, max(0.0, score))


# -----------------------------
# 공식 홈페이지 길드원 수집
# -----------------------------
def make_session():
    session = requests.Session()
    session.headers.update(HEADERS)

    retry = Retry(
        total=5,
        connect=5,
        read=5,
        backoff_factor=1.0,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET"],
    )
    adapter = HTTPAdapter(max_retries=retry)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    return session


def build_search_url(world_name: str, guild_name: str) -> str:
    base_url = WORLD_BASE_URLS.get(world_name)

    if world_name not in WORLD_BASE_URLS:
        raise ValueError(f"지원하지 않는 월드입니다: {world_name}")

    if not base_url:
        raise ValueError(f"{world_name} 월드 URL이 비어 있습니다.")

    return set_query_params(base_url, n=guild_name)


def get_html(session: requests.Session, url: str) -> str:
    response = session.get(url, timeout=20)
    response.raise_for_status()
    return response.text


def extract_guild_detail_url(search_html: str, guild_name: str):
    soup = BeautifulSoup(search_html, "lxml")
    guild_name_norm = normalize_key(guild_name)

    candidates = []

    for a_tag in soup.find_all("a", href=True):
        href = a_tag["href"]
        text = clean_text(a_tag.get_text(" ", strip=True))

        if "/Common/Guild?" not in href:
            continue

        full_url = urljoin(BASE, href)
        candidates.append((text, full_url))

    for text, full_url in candidates:
        if normalize_key(text) == guild_name_norm:
            return full_url

    for text, full_url in candidates:
        if guild_name_norm in normalize_key(text):
            return full_url

    return None


def extract_member_count(page_text: str) -> int:
    match = re.search(r"길드원 수\s*([0-9,]+)명", page_text)
    if not match:
        return 0
    return int(match.group(1).replace(",", ""))


def is_possible_character_name(text: str) -> bool:
    if not text:
        return False

    if len(text) > 20:
        return False

    blocked_prefixes = [
        "길드원 수",
        "주간 명성치",
        "플래그 레이스",
        "지하 수로",
        "길드 마스터",
        "월드 ",
        "레벨",
        "경험치",
        "인기도",
        "직업",
        "길드원 정보",
        "랭킹",
        "Image",
        "Lv.",
        "직위순",
        "레벨순",
    ]

    for prefix in blocked_prefixes:
        if text.startswith(prefix):
            return False

    return True


def parse_member_names_from_detail_html(detail_html: str):
    soup = BeautifulSoup(detail_html, "lxml")
    raw_text = soup.get_text("\n", strip=True)

    lines = [line.strip() for line in raw_text.splitlines()]
    lines = [line for line in lines if line]

    names = []
    i = 0

    while i < len(lines):
        line = lines[i]

        if line in ROLE_WORDS:
            if i + 1 < len(lines):
                candidate = lines[i + 1].strip()
                if is_possible_character_name(candidate):
                    names.append(candidate)
            i += 1
        else:
            i += 1

    unique_names = []
    seen = set()

    for name in names:
        if name not in seen:
            seen.add(name)
            unique_names.append(name)

    return unique_names


def scrape_all_member_names(session: requests.Session, detail_url: str, progress_callback=None):
    sorted_first_url = set_query_params(detail_url, orderby=1, page=1)

    first_html = get_html(session, sorted_first_url)
    first_soup = BeautifulSoup(first_html, "lxml")
    first_text = first_soup.get_text("\n", strip=True)

    member_count = extract_member_count(first_text)
    if member_count <= 0:
        raise RuntimeError("길드원 수를 읽지 못했습니다. 페이지 구조가 바뀌었을 수 있습니다.")

    pages = math.ceil(member_count / 20)
    all_names = []
    seen = set()

    if progress_callback:
        progress_callback("직위순(orderby=1)으로 길드원 수집 시작")
        progress_callback(f"길드원 수: {member_count}명")
        progress_callback(f"총 페이지 수: {pages}")

    for page in range(1, pages + 1):
        page_url = set_query_params(detail_url, orderby=1, page=page)
        page_html = get_html(session, page_url)
        names = parse_member_names_from_detail_html(page_html)

        for name in names:
            if name not in seen:
                seen.add(name)
                all_names.append(name)

        if progress_callback:
            progress_callback(f"{page}/{pages} 페이지 완료, 현재 {len(all_names)}명")

    return all_names


# -----------------------------
# 이미지 / OCR 유틸
# -----------------------------
def read_image_unicode(file_path):
    try:
        data = np.fromfile(file_path, dtype=np.uint8)
        img = cv2.imdecode(data, cv2.IMREAD_COLOR)
        return img
    except Exception:
        return None


def to_bgr(img):
    if img is None:
        return None
    if len(img.shape) == 2:
        return cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)
    return img


def safe_json_loads(value):
    if isinstance(value, str):
        try:
            return json.loads(value)
        except Exception:
            return None
    return value


def create_korean_ocr_engine(log_fn=None):
    attempts = [
        {
            "lang": "korean",
            "text_recognition_model_name": KOREAN_OCR_MODEL_NAME,
            "use_doc_orientation_classify": False,
            "use_doc_unwarping": False,
            "use_textline_orientation": False,
            "show_log": False,
        },
        {
            "lang": "korean",
            "text_recognition_model_name": KOREAN_OCR_MODEL_NAME,
            "show_log": False,
        },
        {
            "lang": "korean",
            "show_log": False,
        },
        {
            "lang": "korean",
        },
    ]

    errors = []

    for kwargs in attempts:
        try:
            engine = PaddleOCR(**kwargs)
            model_name = kwargs.get("text_recognition_model_name", "lang=korean default")
            if log_fn is not None:
                log_fn(f"OCR 엔진 준비 완료 ({model_name})")
            return engine
        except Exception as e:
            errors.append(f"{kwargs}: {e}")

    raise RuntimeError("OCR 엔진 초기화에 실패했습니다. " + " | ".join(errors))


def _ocr_predict_with_method(method, image):
    try:
        return method(image)
    except Exception:
        temp_path = os.path.join(tempfile.gettempdir(), "ocr_temp_input.png")
        ok, encoded = cv2.imencode(".png", image)
        if not ok:
            raise RuntimeError("임시 이미지 저장에 실패했습니다.")
        encoded.tofile(temp_path)
        return method(temp_path)


def ocr_predict_any(ocr_engine, image):
    if hasattr(ocr_engine, "predict"):
        return _ocr_predict_with_method(ocr_engine.predict, image)

    if hasattr(ocr_engine, "ocr"):
        def legacy_method(inp):
            try:
                return ocr_engine.ocr(inp, cls=False)
            except TypeError:
                return ocr_engine.ocr(inp)

        return _ocr_predict_with_method(legacy_method, image)

    raise RuntimeError("지원하지 않는 OCR 엔진입니다. predict/ocr 메서드를 찾지 못했습니다.")


def _collect_legacy_ocr_items(node, items):
    if node is None:
        return

    if isinstance(node, tuple) and len(node) >= 1 and isinstance(node[0], str):
        text = str(node[0]).strip()
        if text:
            score = 0.0
            if len(node) >= 2:
                try:
                    score = float(node[1])
                except Exception:
                    score = 0.0
            items.append({"text": text, "score": score, "x_left": 0.0})
        return

    if isinstance(node, (list, tuple)):
        if len(node) == 2 and isinstance(node[1], (list, tuple)) and len(node[1]) >= 1 and isinstance(node[1][0], str):
            text = str(node[1][0]).strip()
            if text:
                score = 0.0
                if len(node[1]) >= 2:
                    try:
                        score = float(node[1][1])
                    except Exception:
                        score = 0.0

                x_left = 0.0
                box = node[0]
                try:
                    arr = np.array(box)
                    if arr.ndim == 1 and arr.size >= 4:
                        x_left = float(min(arr[0], arr[2]))
                    elif arr.ndim == 2 and arr.shape[1] >= 2:
                        x_left = float(np.min(arr[:, 0]))
                except Exception:
                    x_left = 0.0

                items.append({"text": text, "score": score, "x_left": x_left})
            return

        for sub in node:
            _collect_legacy_ocr_items(sub, items)


def extract_ocr_items(result):
    items = []

    if not result:
        return items

    for item in result:
        data = None

        if hasattr(item, "json"):
            try:
                json_value = item.json
                data = json_value() if callable(json_value) else json_value
            except Exception:
                data = None

        data = safe_json_loads(data)

        if data is None and isinstance(item, dict):
            data = item

        data = safe_json_loads(data)

        if isinstance(data, dict):
            res = data["res"] if "res" in data and isinstance(data["res"], dict) else data

            texts = res.get("rec_texts", [])
            scores = res.get("rec_scores", [])
            boxes = res.get("rec_boxes", [])
            polys = res.get("dt_polys", [])

            for idx, text in enumerate(texts):
                if text is None:
                    continue

                text = str(text).strip()
                if not text:
                    continue

                score = 0.0
                if idx < len(scores):
                    try:
                        score = float(scores[idx])
                    except Exception:
                        score = 0.0

                x_left = 0.0
                box = None

                if idx < len(boxes):
                    box = boxes[idx]
                elif idx < len(polys):
                    box = polys[idx]

                if box is not None:
                    arr = np.array(box)
                    if arr.ndim == 1 and arr.size >= 4:
                        x_left = float(min(arr[0], arr[2]))
                    elif arr.ndim == 2 and arr.shape[1] >= 2:
                        x_left = float(np.min(arr[:, 0]))

                items.append({
                    "text": text,
                    "score": score,
                    "x_left": x_left,
                })
            continue

        _collect_legacy_ocr_items(item, items)

    items.sort(key=lambda v: v["x_left"])
    return items


def build_crop_variants(crop_img, field_type):
    variants = {}

    h, w = crop_img.shape[:2]

    left = int(w * 0.02)
    right = int(w * 0.98)
    top = int(h * 0.08)
    bottom = int(h * 0.92)
    cropped = crop_img[top:bottom, left:right].copy()

    scale = 4.0 if field_type == "name" else 3.5
    enlarged = cv2.resize(cropped, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
    variants["원본확대"] = enlarged

    gray = cv2.cvtColor(enlarged, cv2.COLOR_BGR2GRAY)
    blur = cv2.GaussianBlur(gray, (0, 0), 2.0)
    sharp = cv2.addWeighted(gray, 1.8, blur, -0.8, 0)
    variants["선명"] = to_bgr(sharp)

    if field_type == "number":
        top_hat_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (31, 11))
        top_hat = cv2.morphologyEx(gray, cv2.MORPH_TOPHAT, top_hat_kernel)
        _, binary = cv2.threshold(top_hat, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        variants["숫자이진화"] = to_bgr(binary)

    return variants


def normalize_name_ocr(text):
    if not text:
        return ""
    text = text.strip()
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[^가-힣A-Za-z0-9]", "", text)
    return text


def normalize_number(text):
    if not text:
        return ""

    replace_map = {
        "O": "0", "o": "0", "D": "0", "Q": "0",
        "I": "1", "l": "1", "|": "1",
        "S": "5", "B": "8",
        ".": ",", "·": ",", "'": ",", "`": ",", " ": "",
    }

    for a, b in replace_map.items():
        text = text.replace(a, b)

    filtered = re.sub(r"[^0-9,]", "", text)
    filtered = re.sub(r",+", ",", filtered).strip(",")

    if re.fullmatch(r"\d{1,3}(,\d{3})*", filtered):
        return filtered

    digits = re.findall(r"\d", filtered)
    if not digits:
        return ""

    num = int("".join(digits))
    return f"{num:,}"


def classify_flag_crop(crop_img):
    h, w = crop_img.shape[:2]
    if h == 0 or w == 0:
        return {"text": "", "score": 0.0, "raw_text": "", "variant": "전용판별"}

    left = int(w * 0.03)
    right = max(left + 1, int(w * 0.97))
    top = int(h * 0.08)
    bottom = max(top + 1, int(h * 0.92))
    cropped = crop_img[top:bottom, left:right].copy()
    enlarged = cv2.resize(cropped, None, fx=4.0, fy=4.0, interpolation=cv2.INTER_CUBIC)

    gray = cv2.cvtColor(enlarged, cv2.COLOR_BGR2GRAY)
    top_hat_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (25, 9))
    top_hat = cv2.morphologyEx(gray, cv2.MORPH_TOPHAT, top_hat_kernel)
    _, binary = cv2.threshold(top_hat, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

    ys, xs = np.where(binary > 0)
    if len(xs) == 0 or len(ys) == 0:
        return {"text": "0", "score": 0.60, "raw_text": "blank->0", "variant": "전용판별"}

    x1 = int(xs.min())
    x2 = int(xs.max())
    y1 = int(ys.min())
    y2 = int(ys.max())

    bbox_w = max(1, x2 - x1 + 1)
    bbox_h = max(1, y2 - y1 + 1)
    span_ratio = bbox_w / float(binary.shape[1])

    bbox = binary[y1:y2 + 1, x1:x2 + 1]
    num_labels, labels, stats, _ = cv2.connectedComponentsWithStats(bbox, connectivity=8)
    min_area = max(6, int(round(bbox.size * 0.002)))
    comp_count = 0
    for idx in range(1, num_labels):
        area = int(stats[idx, cv2.CC_STAT_AREA])
        if area >= min_area:
            comp_count += 1

    col_active = (bbox > 0).sum(axis=0) > 0
    x_runs = merge_close_runs(find_runs(col_active), max_gap=max(1, bbox.shape[1] // 45))
    run_count = len(x_runs)

    score_1000 = 0.0
    score_1000 += min(1.0, span_ratio / 0.42) * 0.45
    score_1000 += min(1.0, run_count / 4.0) * 0.35
    score_1000 += min(1.0, comp_count / 4.0) * 0.20

    if run_count >= 3 or comp_count >= 3 or span_ratio >= 0.40:
        text = "1,000"
        score = max(0.65, min(0.99, score_1000))
    else:
        text = "0"
        score = max(0.65, min(0.99, 1.0 - score_1000 * 0.7))

    raw_text = f"span={span_ratio:.3f};runs={run_count};cc={comp_count};bbox={bbox_w}x{bbox_h}"
    return {"text": text, "score": score, "raw_text": raw_text, "variant": "전용판별"}


def normalize_flag(text):
    if not text:
        return ""

    replace_map = {
        "O": "0", "o": "0", "D": "0", "Q": "0",
        "I": "1", "l": "1", "|": "1",
        "S": "5", "B": "8",
    }

    for a, b in replace_map.items():
        text = text.replace(a, b)

    digits = "".join(re.findall(r"\d", text))
    raw = text.replace(" ", "")

    if "1000" in digits:
        return "1,000"
    if "," in raw and len(digits) >= 3:
        return "1,000"
    if len(digits) >= 3 and digits.endswith("000"):
        return "1,000"
    if digits and int(digits) >= 500:
        return "1,000"

    return "0"


def excel_numeric_value(text):
    """엑셀에는 쉼표 없는 숫자만 저장. 숫자가 없으면 빈칸."""
    if text is None:
        return ""
    digits = "".join(re.findall(r"\d", str(text)))
    if not digits:
        return ""
    try:
        return int(digits)
    except Exception:
        return digits


def normalize_text_by_field(text, field_type):
    if field_type == "name":
        return normalize_name_ocr(text)
    if field_type == "number":
        return normalize_number(text)
    if field_type == "flag":
        return normalize_flag(text)
    return text


def score_candidate(text, confidence, field_type):
    if field_type == "name":
        if not text:
            return -999
        penalty = 0
        if text.isdigit():
            penalty += 5
        if len(text) == 1:
            penalty += 3
        return len(text) * 3 + confidence * 4 - penalty

    if field_type == "number":
        if not text:
            return -999
        digit_count = len(re.findall(r"\d", text))
        return digit_count * 3 + confidence * 4

    if field_type == "flag":
        if text not in ("0", "1,000"):
            return -999
        return 20 + confidence * 4

    return confidence


def recognize_one_crop(crop_img, field_type, ocr_engine):
    variants = build_crop_variants(crop_img, field_type)
    candidates = []

    for variant_name, variant_img in variants.items():
        result = ocr_predict_any(ocr_engine, variant_img)
        items = extract_ocr_items(result)

        if items:
            raw_text = "".join(item["text"] for item in items)
            avg_score = sum(item["score"] for item in items) / len(items)
        else:
            raw_text = ""
            avg_score = 0.0

        normalized = normalize_text_by_field(raw_text, field_type)

        candidates.append({
            "text": normalized,
            "score": avg_score,
            "raw_text": raw_text,
            "variant": variant_name,
        })

    best = max(candidates, key=lambda c: score_candidate(c["text"], c["score"], field_type))
    return best


# -----------------------------
# 점수 열(F/G) 기반 row segmentation
# -----------------------------
def smooth_1d(arr: np.ndarray, window: int) -> np.ndarray:
    window = max(1, int(window))
    if window <= 1:
        return arr.copy()

    kernel = np.ones(window, dtype=np.float32) / window
    return np.convolve(arr, kernel, mode="same")


def ensure_odd(n: int) -> int:
    return n if n % 2 == 1 else n + 1


def compute_row_text_features(gray: np.ndarray) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    h, w = gray.shape

    x1 = int(w * 0.02)
    x2 = int(w * 0.98)
    roi = gray[:, x1:x2]

    kx = ensure_odd(max(15, int(w * 0.08)))
    ky = ensure_odd(max(9, int(h * 0.03)))
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (kx, ky))
    top_hat = cv2.morphologyEx(roi, cv2.MORPH_TOPHAT, kernel)

    p90 = float(np.percentile(top_hat, 90))
    p99 = float(np.percentile(top_hat, 99.5))

    if p99 <= p90:
        cut = p90
    else:
        cut = p90 + 0.35 * (p99 - p90)

    bright_excess = np.clip(top_hat.astype(np.float32) - cut, 0, None)
    row_scores = bright_excess.mean(axis=1)
    row_scores = smooth_1d(row_scores, max(5, h // 120))

    active_mask = top_hat > cut
    active_counts = active_mask.sum(axis=1).astype(np.int32)
    return row_scores, active_counts, active_mask


def find_text_band(row_scores: np.ndarray) -> tuple[int, int]:
    max_score = float(np.max(row_scores))
    if max_score <= 0:
        return 0, len(row_scores) - 1

    band_threshold = max_score * 0.18
    active = np.where(row_scores >= band_threshold)[0]

    if len(active) == 0:
        return 0, len(row_scores) - 1

    top = int(active[0])
    bottom = int(active[-1])

    pad = max(2, len(row_scores) // 200)
    top = max(0, top - pad)
    bottom = min(len(row_scores) - 1, bottom + pad)

    return top, bottom


def refine_centers_simple(row_scores: np.ndarray, top: int, bottom: int, row_count: int) -> list[int]:
    band_height = bottom - top + 1
    step = band_height / row_count

    initial_centers = []
    for i in range(row_count):
        c = top + (i + 0.5) * step
        initial_centers.append(c)

    refined = []
    search_radius = max(3, int(step * 0.35))

    for c in initial_centers:
        center = int(round(c))
        lo = max(top, center - search_radius)
        hi = min(bottom, center + search_radius)

        if lo >= hi:
            refined.append(center)
            continue

        local_idx = int(np.argmax(row_scores[lo:hi + 1]))
        best_y = lo + local_idx
        refined.append(best_y)

    min_sep = max(3, int(step * 0.45))

    for i in range(1, len(refined)):
        if refined[i] <= refined[i - 1] + min_sep:
            refined[i] = refined[i - 1] + min_sep

    for i in range(len(refined) - 2, -1, -1):
        if refined[i] >= refined[i + 1] - min_sep:
            refined[i] = refined[i + 1] - min_sep

    refined = [max(top, min(bottom, y)) for y in refined]
    return refined


def get_min_blank_run_length(start_y: int, end_y: int) -> int:
    gap = max(1, int(end_y) - int(start_y))
    return max(2, min(6, int(round(gap * 0.06))))


def compute_eroded_active_counts(active_mask: np.ndarray) -> np.ndarray:
    mask_u8 = active_mask.astype(np.uint8) * 255
    erosion_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 3))
    eroded_mask = cv2.erode(mask_u8, erosion_kernel, iterations=1)
    return (eroded_mask > 0).sum(axis=1).astype(np.int32)


def find_blank_cut_in_range(active_counts: np.ndarray, start_y: int, end_y: int, min_run_length: int = 2):
    start_y = max(0, int(start_y))
    end_y = min(len(active_counts) - 1, int(end_y))
    min_run_length = max(1, int(min_run_length))

    if end_y < start_y:
        return None

    blank_mask = active_counts[start_y:end_y + 1] == 0
    runs = find_runs(blank_mask)
    runs = [run for run in runs if (run[1] - run[0] + 1) >= min_run_length]

    if not runs:
        return None

    target_mid = (end_y - start_y) / 2.0
    best_run = max(
        runs,
        key=lambda run: (
            run[1] - run[0] + 1,
            -abs(((run[0] + run[1]) / 2.0) - target_mid),
        ),
    )
    return start_y + (best_run[0] + best_run[1]) // 2


def find_blank_cut_with_fallback(
    active_counts: np.ndarray,
    eroded_active_counts: np.ndarray | None,
    start_y: int,
    end_y: int,
    min_run_length: int = 2,
):
    direct_cut = find_blank_cut_in_range(
        active_counts,
        start_y,
        end_y,
        min_run_length=min_run_length,
    )
    if direct_cut is not None:
        return direct_cut

    if eroded_active_counts is not None:
        eroded_cut = find_blank_cut_in_range(
            eroded_active_counts,
            start_y,
            end_y,
            min_run_length=min_run_length,
        )
        if eroded_cut is not None:
            return eroded_cut

    return None


def centers_to_bounds(
    centers: list[int],
    image_height: int,
    active_counts: np.ndarray | None = None,
    eroded_active_counts: np.ndarray | None = None,
) -> list[tuple[int, int]]:
    bounds = []

    if not centers:
        return bounds

    row_count = len(centers)

    default_mids = []
    for i in range(row_count - 1):
        default_mids.append((centers[i] + centers[i + 1]) // 2)

    first_half = max(1, centers[1] - centers[0]) if row_count >= 2 else max(10, image_height // 20)
    last_half = max(1, centers[-1] - centers[-2]) if row_count >= 2 else max(10, image_height // 20)

    top_y = max(0, centers[0] - first_half // 2)
    bottom_y = min(image_height - 1, centers[-1] + last_half // 2)

    current_top = top_y
    for i in range(row_count):
        if i < row_count - 1:
            current_bottom = default_mids[i]

            if active_counts is not None:
                min_blank_run = get_min_blank_run_length(centers[i], centers[i + 1])
                blank_cut = find_blank_cut_with_fallback(
                    active_counts,
                    eroded_active_counts,
                    centers[i],
                    centers[i + 1],
                    min_run_length=min_blank_run,
                )
                if blank_cut is not None:
                    current_bottom = blank_cut
        else:
            current_bottom = bottom_y

        current_bottom = max(current_top, current_bottom)
        bounds.append((int(current_top), int(current_bottom)))
        current_top = current_bottom + 1

    fixed = []
    for y1, y2 in bounds:
        y1 = max(0, min(image_height - 1, y1))
        y2 = max(0, min(image_height - 1, y2))
        if y2 < y1:
            y2 = y1
        fixed.append((y1, y2))

    return fixed


def get_row_bounds_from_nickname_image(nickname_img: np.ndarray, row_count: int) -> list[tuple[int, int]]:
    gray = cv2.cvtColor(nickname_img, cv2.COLOR_BGR2GRAY)
    row_scores, active_counts, active_mask = compute_row_text_features(gray)
    eroded_active_counts = compute_eroded_active_counts(active_mask)
    top, bottom = find_text_band(row_scores)
    centers = refine_centers_simple(row_scores, top, bottom, row_count)
    bounds = centers_to_bounds(
        centers,
        gray.shape[0],
        active_counts=active_counts,
        eroded_active_counts=eroded_active_counts,
    )
    return bounds


def build_reference_descriptor(column_img: np.ndarray, row_count: int, column_name: str) -> dict:
    gray = cv2.cvtColor(column_img, cv2.COLOR_BGR2GRAY)
    row_scores, active_counts, active_mask = compute_row_text_features(gray)
    top, bottom = find_text_band(row_scores)
    centers = refine_centers_simple(row_scores, top, bottom, row_count)

    sampled_active = []
    sampled_scores = []
    for center in centers:
        lo = max(0, center - 1)
        hi = min(len(active_counts) - 1, center + 1)
        sampled_active.append(int(np.max(active_counts[lo:hi + 1])))
        sampled_scores.append(float(np.max(row_scores[lo:hi + 1])))

    median_active = float(np.median(sampled_active)) if sampled_active else 0.0
    median_score = float(np.median(sampled_scores)) if sampled_scores else 0.0
    band_height = float(max(0, bottom - top + 1))
    positive_density = float(np.mean(active_counts[top:bottom + 1] > 0)) if bottom >= top else 0.0

    quality = (median_active * 4.0) + (median_score * 40.0) + (band_height * 0.35) + (positive_density * 25.0)

    return {
        "name": column_name,
        "gray": gray,
        "row_scores": row_scores,
        "active_counts": active_counts,
        "active_mask": active_mask,
        "top": top,
        "bottom": bottom,
        "centers": centers,
        "quality": quality,
    }


def find_local_active_run(
    active_counts: np.ndarray,
    start_y: int,
    end_y: int,
    center_y: int,
    min_run_length: int = 1,
):
    start_y = max(0, int(start_y))
    end_y = min(len(active_counts) - 1, int(end_y))
    center_y = max(start_y, min(end_y, int(center_y)))

    if end_y < start_y:
        return None

    local_counts = active_counts[start_y:end_y + 1]
    positive = local_counts[local_counts > 0]
    if len(positive) == 0:
        return None

    base_threshold = np.percentile(positive, 25)
    min_active = max(1, int(round(base_threshold * 0.35)))
    active_mask = local_counts >= min_active
    runs = find_runs(active_mask)
    runs = [run for run in runs if (run[1] - run[0] + 1) >= max(1, int(min_run_length))]

    if not runs:
        return None

    center_rel = center_y - start_y
    containing = [run for run in runs if run[0] <= center_rel <= run[1]]
    candidates = containing if containing else runs

    def rank(run):
        rs, re = run
        height = re - rs + 1
        strength = int(local_counts[rs:re + 1].sum())
        midpoint = (rs + re) / 2.0
        distance = abs(midpoint - center_rel)
        return (
            -distance,
            height,
            strength,
        )

    best = max(candidates, key=rank)
    rs, re = best
    strength = int(local_counts[rs:re + 1].sum())

    return {
        "start": start_y + rs,
        "end": start_y + re,
        "height": re - rs + 1,
        "strength": strength,
        "threshold": min_active,
    }


def choose_better_local_run(run_a, run_b):
    if run_a is None:
        return run_b
    if run_b is None:
        return run_a

    score_a = (run_a["height"] * 1000) + run_a["strength"]
    score_b = (run_b["height"] * 1000) + run_b["strength"]
    return run_a if score_a >= score_b else run_b


def combine_descriptor_active_counts(descriptors: list[dict]) -> tuple[np.ndarray, np.ndarray | None]:
    active_counts = [desc["active_counts"] for desc in descriptors if "active_counts" in desc]
    eroded_counts_list = []

    if not active_counts:
        raise RuntimeError("점수 열 active_counts를 결합할 수 없습니다.")

    combined_counts = np.maximum.reduce(active_counts)

    for desc in descriptors:
        active_mask = desc.get("active_mask")
        if active_mask is None:
            continue
        # F/G 열은 폭이 서로 다를 수 있으므로 2D mask 자체를 바로 OR-reduce 하지 않는다.
        # 각 열에서 erosion한 뒤, 1D row counts로 변환해서 결합한다.
        eroded_counts_list.append(compute_eroded_active_counts(active_mask))

    if eroded_counts_list:
        eroded_active_counts = np.maximum.reduce(eroded_counts_list)
    else:
        eroded_active_counts = None

    return combined_counts, eroded_active_counts


def build_bounds_from_score_reference(descriptors: list[dict], base_centers: list[int], image_height: int) -> list[tuple[int, int]]:
    row_count = len(base_centers)

    if row_count == 0:
        return []

    combined_active_counts, combined_eroded_counts = combine_descriptor_active_counts(descriptors)

    # row 중심은 F/G 중 품질이 더 좋은 열에서 잡되,
    # 실제 경계는 F/G 결합 active profile에서 잘라서 절대 겹치지 않게 만든다.
    bounds = centers_to_bounds(
        base_centers,
        image_height,
        active_counts=combined_active_counts,
        eroded_active_counts=combined_eroded_counts,
    )

    fixed = []
    prev_bottom = -1
    for y1, y2 in bounds:
        y1 = max(prev_bottom + 1, int(y1))
        y2 = max(y1, int(y2))
        y2 = min(image_height - 1, y2)
        fixed.append((y1, y2))
        prev_bottom = y2

    return fixed


def get_row_bounds_from_score_columns(suro_img: np.ndarray, flag_img: np.ndarray, row_count: int) -> list[tuple[int, int]]:
    descriptors = [
        build_reference_descriptor(suro_img, row_count, "suro"),
        build_reference_descriptor(flag_img, row_count, "flag"),
    ]

    base_descriptor = max(descriptors, key=lambda d: d["quality"])
    centers = base_descriptor["centers"]
    bounds = build_bounds_from_score_reference(
        descriptors,
        centers,
        base_descriptor["gray"].shape[0],
    )
    return bounds


def normalize_profile_1d(values: np.ndarray) -> np.ndarray:
    arr = np.asarray(values, dtype=np.float32).reshape(-1)
    if arr.size == 0:
        return arr
    finite = arr[np.isfinite(arr)]
    if finite.size == 0:
        return np.zeros_like(arr, dtype=np.float32)
    vmax = float(np.max(finite))
    if vmax <= 1e-6:
        return np.zeros_like(arr, dtype=np.float32)
    return np.clip(arr / vmax, 0.0, 1.0).astype(np.float32)


def build_combined_row_profile(nickname_img: np.ndarray, suro_img: np.ndarray, flag_img: np.ndarray):
    imgs = {
        "nickname": nickname_img,
        "suro": suro_img,
        "flag": flag_img,
    }

    features = {}
    for key, img in imgs.items():
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        row_scores, active_counts, active_mask = compute_row_text_features(gray)
        features[key] = {
            "gray": gray,
            "row_scores": row_scores,
            "active_counts": active_counts,
            "active_mask": active_mask,
        }

    nick_score = normalize_profile_1d(features["nickname"]["row_scores"])
    suro_score = normalize_profile_1d(features["suro"]["row_scores"])
    flag_score = normalize_profile_1d(features["flag"]["row_scores"])

    nick_active = normalize_profile_1d(features["nickname"]["active_counts"].astype(np.float32))
    suro_active = normalize_profile_1d(features["suro"]["active_counts"].astype(np.float32))
    flag_active = normalize_profile_1d(features["flag"]["active_counts"].astype(np.float32))

    combined = (
        0.42 * nick_score
        + 0.22 * nick_active
        + 0.20 * suro_score
        + 0.08 * suro_active
        + 0.06 * flag_score
        + 0.02 * flag_active
    ).astype(np.float32)

    combined = smooth_1d(combined, max(7, len(combined) // 120))

    return {
        "combined_scores": combined,
        "nickname": features["nickname"],
        "suro": features["suro"],
        "flag": features["flag"],
    }


def refine_centers_on_global_grid(row_scores: np.ndarray, top: int, bottom: int, row_count: int) -> list[int]:
    if row_count <= 0:
        return []

    band_height = max(1, bottom - top + 1)
    step = band_height / float(row_count)

    centers = []
    for i in range(row_count):
        seg_lo = int(round(top + (i * step)))
        seg_hi = int(round(top + ((i + 1) * step))) - 1

        if i == row_count - 1:
            seg_hi = bottom

        seg_lo = max(top, min(bottom, seg_lo))
        seg_hi = max(seg_lo, min(bottom, seg_hi))

        segment = row_scores[seg_lo:seg_hi + 1]
        if segment.size == 0:
            center = int(round(top + ((i + 0.5) * step)))
            center = max(top, min(bottom, center))
        else:
            local_max = float(np.max(segment))
            if local_max <= 1e-6:
                center = int(round((seg_lo + seg_hi) / 2.0))
            else:
                center = seg_lo + int(np.argmax(segment))

        if centers and center <= centers[-1]:
            center = min(bottom, max(centers[-1] + 1, center))
        centers.append(center)

    for i in range(len(centers) - 2, -1, -1):
        if centers[i] >= centers[i + 1]:
            centers[i] = max(top, centers[i + 1] - 1)

    return [int(max(top, min(bottom, c))) for c in centers]


def centers_to_bounds_midpoint(centers: list[int], top: int, bottom: int, image_height: int) -> list[tuple[int, int]]:
    if not centers:
        return []

    top = max(0, min(image_height - 1, int(top)))
    bottom = max(0, min(image_height - 1, int(bottom)))
    if bottom < top:
        bottom = top

    bounds = []
    current_top = top
    for i, center in enumerate(centers):
        if i < len(centers) - 1:
            current_bottom = (centers[i] + centers[i + 1]) // 2
        else:
            current_bottom = bottom

        current_bottom = max(current_top, min(image_height - 1, int(current_bottom)))
        bounds.append((int(current_top), int(current_bottom)))
        current_top = min(image_height - 1, current_bottom + 1)

    fixed = []
    for y1, y2 in bounds:
        y1 = max(0, min(image_height - 1, int(y1)))
        y2 = max(0, min(image_height - 1, int(y2)))
        if y2 < y1:
            y2 = y1
        fixed.append((y1, y2))

    return fixed


def get_row_bounds_global_grid(nickname_img: np.ndarray, suro_img: np.ndarray, flag_img: np.ndarray, row_count: int) -> list[tuple[int, int]]:
    profile_info = build_combined_row_profile(nickname_img, suro_img, flag_img)
    combined_scores = profile_info["combined_scores"]

    top, bottom = find_text_band(combined_scores)

    band_height = max(1, bottom - top + 1)
    step = band_height / float(max(1, row_count))
    outer_pad = max(2, int(round(step * 0.18)))
    top = max(0, top - outer_pad)
    bottom = min(len(combined_scores) - 1, bottom + outer_pad)

    centers = refine_centers_on_global_grid(combined_scores, top, bottom, row_count)
    bounds = centers_to_bounds_midpoint(centers, top, bottom, len(combined_scores))
    return bounds


def scale_bounds_to_target(bounds, src_height, dst_height):
    scaled = []

    for y1, y2 in bounds:
        ny1 = int(round((y1 / src_height) * dst_height))
        ny2 = int(round((y2 / src_height) * dst_height))

        ny1 = max(0, min(dst_height - 1, ny1))
        ny2 = max(0, min(dst_height - 1, ny2))
        if ny2 < ny1:
            ny2 = ny1

        scaled.append((ny1, ny2))

    return scaled


def crop_with_bounds(image: np.ndarray, bounds: list[tuple[int, int]]) -> list[np.ndarray]:
    rows = []
    h, w = image.shape[:2]

    for y1, y2 in bounds:
        y1 = max(0, min(h - 1, y1))
        y2 = max(0, min(h - 1, y2))
        if y2 < y1:
            y2 = y1
        rows.append(image[y1:y2 + 1, 0:w].copy())

    return rows


def image_to_png_bytes(image: np.ndarray) -> bytes:
    ok, encoded = cv2.imencode(".png", image)
    if not ok:
        return b""
    return encoded.tobytes()


# -----------------------------
# 원본 이미지에서 열 자동 검출
# -----------------------------
def save_image_unicode(file_path, image):
    ok, enc = cv2.imencode(".png", image)
    if not ok:
        return False
    enc.tofile(file_path)
    return True


def find_runs(mask: np.ndarray):
    runs = []
    start = None

    for i, v in enumerate(mask):
        if v and start is None:
            start = i
        elif not v and start is not None:
            runs.append((start, i - 1))
            start = None

    if start is not None:
        runs.append((start, len(mask) - 1))

    return runs


def merge_close_runs(runs, max_gap=10):
    if not runs:
        return []

    merged = [list(runs[0])]

    for s, e in runs[1:]:
        prev_s, prev_e = merged[-1]
        if s - prev_e <= max_gap:
            merged[-1][1] = e
        else:
            merged.append([s, e])

    return [(s, e) for s, e in merged]


def detect_columns_from_full_image(image):
    h, w = image.shape[:2]
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    y1 = int(h * 0.02)
    y2 = int(h * 0.98)
    roi = gray[y1:y2, :]

    kx = ensure_odd(max(21, int(w * 0.03)))
    ky = ensure_odd(max(9, int(h * 0.06)))
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (kx, ky))
    top_hat = cv2.morphologyEx(roi, cv2.MORPH_TOPHAT, kernel)

    p90 = float(np.percentile(top_hat, 90))
    p99 = float(np.percentile(top_hat, 99.5))
    if p99 <= p90:
        cut = p90
    else:
        cut = p90 + 0.30 * (p99 - p90)

    bright_excess = np.clip(top_hat.astype(np.float32) - cut, 0, None)

    col_scores = bright_excess.mean(axis=0)
    col_scores = smooth_1d(col_scores, max(9, w // 120))

    max_score = float(np.max(col_scores))
    if max_score <= 0:
        raise RuntimeError("텍스트 열을 찾지 못했습니다.")

    active_threshold = max_score * 0.18
    active = col_scores >= active_threshold
    runs = find_runs(active)
    runs = merge_close_runs(runs, max_gap=max(6, w // 150))

    min_width = max(8, int(w * 0.015))
    runs = [(s, e) for s, e in runs if (e - s + 1) >= min_width]

    if len(runs) < 3:
        raise RuntimeError(f"필요한 열을 충분히 찾지 못했습니다. 감지된 열 개수: {len(runs)}")

    nickname_run = runs[0]
    right_two = runs[-2:]
    suro_run = right_two[0]
    flag_run = right_two[1]

    def pad_run(run, left_pad_ratio, right_pad_ratio):
        s, e = run
        pad_l = max(3, int(w * left_pad_ratio))
        pad_r = max(3, int(w * right_pad_ratio))
        s = max(0, s - pad_l)
        e = min(w - 1, e + pad_r)
        return (s, e)

    nickname_run = pad_run(nickname_run, 0.012, 0.055)
    # F열은 중앙 정렬이고 최대 10만 단위까지 갈 수 있어 좌우를 넉넉하게 확장
    # 기존 0.015/0.015 대비 2배인 0.030/0.030 적용
    suro_run = pad_run(suro_run, 0.030, 0.030)
    flag_run = pad_run(flag_run, 0.012, 0.012)

    return {
        "nickname": nickname_run,
        "suro": suro_run,
        "flag": flag_run,
        "all_runs": runs,
        "col_scores": col_scores,
    }


def crop_columns(image, detected):
    h, w = image.shape[:2]

    def crop_run(run):
        x1, x2 = run
        x1 = max(0, min(w - 1, x1))
        x2 = max(0, min(w - 1, x2))
        if x2 < x1:
            x2 = x1
        return image[:, x1:x2 + 1].copy()

    nickname_img = crop_run(detected["nickname"])
    suro_img = crop_run(detected["suro"])
    flag_img = crop_run(detected["flag"])

    return nickname_img, suro_img, flag_img


# -----------------------------
# 하이브리드 닉네임 매칭
#   - 0.75 이상 고신뢰는 1회 사용으로 잠금
#   - 그 미만은 중복 허용하되, 중복 패널티로 최대한 분산
# -----------------------------
def run_global_unique_matching(accumulated_rows, official_names):
    rows = [dict(row) for row in accumulated_rows]

    if not rows or not official_names:
        return rows

    m = len(rows)
    n = len(official_names)

    score_matrix = np.zeros((m, n), dtype=np.float64)
    ranked_candidates = []

    for i, row in enumerate(rows):
        ocr_name = row.get("ocr_nickname", "")
        per_row = []

        for j, official_name in enumerate(official_names):
            score = mixed_name_similarity(ocr_name, official_name)
            score_matrix[i, j] = score
            per_row.append((j, official_name, float(score)))

        per_row.sort(key=lambda x: x[2], reverse=True)
        ranked_candidates.append(per_row)

    matched = {}
    locked_rows = set()
    locked_names = set()

    # 1) 고신뢰 후보는 강하게 잠금 (중복 불가)
    high_conf_pairs = []
    for i, candidates in enumerate(ranked_candidates):
        for j, official_name, score in candidates:
            if score >= HIGH_CONFIDENCE_LOCK_THRESHOLD:
                high_conf_pairs.append((score, i, j, official_name))

    high_conf_pairs.sort(key=lambda x: (-x[0], x[1], x[2]))

    for score, row_idx, col_idx, official_name in high_conf_pairs:
        if row_idx in locked_rows:
            continue
        if official_name in locked_names:
            continue

        matched[row_idx] = (official_name, float(score))
        locked_rows.add(row_idx)
        locked_names.add(official_name)

    # 2) 나머지는 행별 최고 후보를 우선하되, 중복이 많아질수록 패널티
    usage_counts = {name: 0 for name in official_names}
    for official_name, _score in matched.values():
        if official_name:
            usage_counts[official_name] = usage_counts.get(official_name, 0) + 1

    remaining_rows = [i for i in range(m) if i not in locked_rows]
    remaining_rows.sort(
        key=lambda row_idx: ranked_candidates[row_idx][0][2] if ranked_candidates[row_idx] else 0.0,
        reverse=True,
    )

    for row_idx in remaining_rows:
        candidates = ranked_candidates[row_idx]
        best_name = ""
        best_score = 0.0
        best_objective = -1e9

        for _col_idx, official_name, score in candidates:
            if official_name in locked_names:
                continue

            duplicate_count = usage_counts.get(official_name, 0)
            objective = score - (DUPLICATE_PENALTY_PER_EXTRA_USE * duplicate_count)

            if objective > best_objective:
                best_objective = objective
                best_name = official_name
                best_score = float(score)

        if not best_name and candidates:
            best_name = candidates[0][1]
            best_score = float(candidates[0][2])

        if best_name:
            matched[row_idx] = (best_name, best_score)
            usage_counts[best_name] = usage_counts.get(best_name, 0) + 1
        else:
            matched[row_idx] = ("", 0.0)

    for i in range(m):
        corrected_name, score = matched.get(i, ("", 0.0))

        if corrected_name:
            rows[i]["corrected_nickname"] = corrected_name
            rows[i]["match_score"] = f"{score:.3f}"
        else:
            rows[i]["corrected_nickname"] = rows[i].get("ocr_nickname", "")
            rows[i]["match_score"] = ""

    return rows



# -----------------------------
# Worker
# -----------------------------
class GuildFetchWorker(QThread):
    log_signal = Signal(str)
    done_signal = Signal(list)
    error_signal = Signal(str)

    def __init__(self, world_name: str, guild_name: str):
        super().__init__()
        self.world_name = world_name
        self.guild_name = guild_name

    def log(self, text: str):
        self.log_signal.emit(text)

    def run(self):
        try:
            self.log("세션 생성 중...")
            session = make_session()

            search_url = build_search_url(self.world_name, self.guild_name)
            self.log("검색 URL 생성 완료")
            self.log(search_url)

            self.log("검색 페이지 불러오는 중...")
            search_html = get_html(session, search_url)

            detail_url = extract_guild_detail_url(search_html, self.guild_name)
            if not detail_url:
                raise RuntimeError("검색 결과에서 길드 상세 링크를 찾지 못했습니다.")

            self.log("길드 상세 페이지 찾음")
            self.log(detail_url)

            member_names = scrape_all_member_names(session, detail_url, self.log)
            if not member_names:
                raise RuntimeError("길드원 이름을 하나도 찾지 못했습니다.")

            self.done_signal.emit(member_names)

        except Exception as e:
            self.error_signal.emit(str(e))


class BatchOCRWorker(QThread):
    log_signal = Signal(str)
    done_signal = Signal(list, list)
    error_signal = Signal(str)

    def __init__(self, batch_items):
        super().__init__()
        self.batch_items = batch_items

    def log(self, text):
        self.log_signal.emit(text)

    def _process_single_image(self, image_path, row_count, ocr_engine):
        self.log("원본 이미지 불러오는 중...")
        full_img = read_image_unicode(image_path)

        if full_img is None:
            raise RuntimeError("원본 이미지를 읽지 못했습니다.")

        self.log("원본 이미지에서 열 자동 추출 중...")
        detected = detect_columns_from_full_image(full_img)
        nickname_img, suro_img, flag_img = crop_columns(full_img, detected)

        self.log("A/F/G 결합 전역 grid 기준으로 row segmentation 중...")
        global_bounds = get_row_bounds_global_grid(nickname_img, suro_img, flag_img, row_count)

        self.log("전역 row 경계를 닉네임/수로/플래그 열에 적용 중...")
        nickname_bounds = scale_bounds_to_target(global_bounds, nickname_img.shape[0], nickname_img.shape[0])
        suro_bounds = scale_bounds_to_target(global_bounds, nickname_img.shape[0], suro_img.shape[0])
        flag_bounds = scale_bounds_to_target(global_bounds, nickname_img.shape[0], flag_img.shape[0])

        nickname_rows = crop_with_bounds(nickname_img, nickname_bounds)
        suro_rows = crop_with_bounds(suro_img, suro_bounds)
        flag_rows = crop_with_bounds(flag_img, flag_bounds)

        merged = []

        for i in range(row_count):
            self.log(f"{i + 1}/{row_count} row 처리 중...")

            nick_res = recognize_one_crop(nickname_rows[i], "name", ocr_engine)
            suro_res = recognize_one_crop(suro_rows[i], "number", ocr_engine)
            flag_res = classify_flag_crop(flag_rows[i])

            merged.append({
                "row": i + 1,
                "local_row": i + 1,
                "source_image_name": os.path.basename(image_path),
                "source_image_path": image_path,
                "ocr_nickname": nick_res["text"],
                "corrected_nickname": nick_res["text"],
                "match_score": "",
                "suro": suro_res["text"],
                "flag": flag_res["text"],
                "nickname_image_bytes": image_to_png_bytes(nickname_rows[i]),
            })

        return merged

    def run(self):
        try:
            if not self.batch_items:
                raise RuntimeError("처리할 이미지가 없습니다.")

            self.log("OCR 엔진 준비 중... (A: korean_PP-OCRv5_mobile_rec 시도 / F: row별 OCR+숫자·쉼표 정규화 / G: 전용 0·1,000 판별)")
            ocr_engine = create_korean_ocr_engine(self.log)

            all_rows = []
            batch_summary = []
            total_images = len(self.batch_items)

            for image_idx, item in enumerate(self.batch_items, start=1):
                image_path = item["path"]
                row_count = item["row_count"]
                image_name = os.path.basename(image_path)

                self.log("=" * 50)
                self.log(f"[{image_idx}/{total_images}] 처리 시작: {image_name} / row {row_count}")

                try:
                    image_rows = self._process_single_image(image_path, row_count, ocr_engine)

                    start_row = len(all_rows) + 1
                    for offset, row in enumerate(image_rows):
                        row["row"] = start_row + offset

                    all_rows.extend(image_rows)
                    batch_summary.append({
                        "image_name": image_name,
                        "image_path": image_path,
                        "row_count": row_count,
                        "produced_rows": len(image_rows),
                        "status": "완료",
                    })
                    self.log(f"[{image_idx}/{total_images}] 완료: {image_name} / {len(image_rows)}개 row")

                except Exception as image_error:
                    error_message = f"{image_name}: {image_error}"
                    batch_summary.append({
                        "image_name": image_name,
                        "image_path": image_path,
                        "row_count": row_count,
                        "produced_rows": 0,
                        "status": f"실패 - {image_error}",
                    })
                    self.log(f"[{image_idx}/{total_images}] 실패: {error_message}")

            if not all_rows:
                raise RuntimeError("모든 이미지 처리에 실패했습니다. 로그를 확인하세요.")

            self.done_signal.emit(all_rows, batch_summary)

        except Exception as e:
            self.error_signal.emit(str(e))


# -----------------------------
# 메인 앱
# -----------------------------
class IntegratedApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("길드원 명단 + OCR 자동보정 통합 앱")
        self.resize(1180, 900)

        self.guild_worker = None
        self.ocr_worker = None

        self.official_names = []
        self.accumulated_rows = []
        self.processed_batches = []

        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()

        layout.addWidget(QLabel("1) 공식 길드원 명단 불러오기"))

        guild_row = QHBoxLayout()
        guild_row.addWidget(QLabel("월드"))

        self.world_combo = QComboBox()
        self.world_combo.addItems(WORLD_BASE_URLS.keys())
        guild_row.addWidget(self.world_combo)

        guild_row.addWidget(QLabel("길드 이름"))
        self.guild_input = QLineEdit()
        self.guild_input.setPlaceholderText("길드 이름 입력")
        guild_row.addWidget(self.guild_input)

        self.fetch_guild_button = QPushButton("공식 길드원 불러오기")
        self.fetch_guild_button.clicked.connect(self.fetch_guild_members)
        guild_row.addWidget(self.fetch_guild_button)

        layout.addLayout(guild_row)

        self.official_list = QListWidget()
        layout.addWidget(self.official_list)

        layout.addWidget(QLabel("2) 원본 이미지 여러 장 등록"))

        batch_control_row = QHBoxLayout()
        batch_control_row.addWidget(QLabel("기본 row 수"))

        self.row_spin = QSpinBox()
        self.row_spin.setMinimum(1)
        self.row_spin.setMaximum(500)
        self.row_spin.setValue(17)
        batch_control_row.addWidget(self.row_spin)

        self.add_images_button = QPushButton("원본 이미지 여러 장 추가")
        self.add_images_button.clicked.connect(self.add_batch_images)
        batch_control_row.addWidget(self.add_images_button)

        self.remove_images_button = QPushButton("선택 이미지 삭제")
        self.remove_images_button.clicked.connect(self.remove_selected_batch_images)
        batch_control_row.addWidget(self.remove_images_button)

        self.clear_images_button = QPushButton("이미지 목록 비우기")
        self.clear_images_button.clicked.connect(self.clear_batch_images)
        batch_control_row.addWidget(self.clear_images_button)

        layout.addLayout(batch_control_row)

        self.batch_summary_label = QLabel("등록 이미지: 0장")
        layout.addWidget(self.batch_summary_label)

        self.batch_table = QTableWidget()
        self.batch_table.setColumnCount(3)
        self.batch_table.setHorizontalHeaderLabels(["파일명", "전체 경로", "row 수"])
        self.batch_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.batch_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.batch_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        layout.addWidget(self.batch_table)

        action_row = QHBoxLayout()

        self.run_ocr_button = QPushButton("목록 전체 OCR 실행")
        self.run_ocr_button.clicked.connect(self.run_ocr)
        action_row.addWidget(self.run_ocr_button)

        self.correct_button = QPushButton("공식 명단으로 자동보정")
        self.correct_button.clicked.connect(self.apply_auto_correction)
        action_row.addWidget(self.correct_button)

        self.reset_button = QPushButton("초기화")
        self.reset_button.clicked.connect(self.reset_all)
        action_row.addWidget(self.reset_button)

        self.save_button = QPushButton("엑셀 저장")
        self.save_button.clicked.connect(self.save_excel)
        action_row.addWidget(self.save_button)

        layout.addLayout(action_row)

        layout.addWidget(QLabel("진행 로그"))
        self.log_box = QTextEdit()
        self.log_box.setReadOnly(True)
        layout.addWidget(self.log_box)

        layout.addWidget(QLabel("3) 결과 표"))
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels([
            "Row", "OCR 닉네임", "보정 닉네임", "매칭 점수", "수로 점수", "플래그 점수"
        ])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.table)

        self.setLayout(layout)

    def log(self, text: str):
        self.log_box.append(text)

    def set_ui_enabled(self, enabled: bool):
        self.world_combo.setEnabled(enabled)
        self.guild_input.setEnabled(enabled)
        self.fetch_guild_button.setEnabled(enabled)
        self.row_spin.setEnabled(enabled)
        self.add_images_button.setEnabled(enabled)
        self.remove_images_button.setEnabled(enabled)
        self.clear_images_button.setEnabled(enabled)
        self.batch_table.setEnabled(enabled)
        self.run_ocr_button.setEnabled(enabled)
        self.correct_button.setEnabled(enabled)
        self.reset_button.setEnabled(enabled)
        self.save_button.setEnabled(enabled)

    def select_images_common(self, title):
        file_paths, _ = QFileDialog.getOpenFileNames(
            self,
            title,
            "",
            "Image Files (*.png *.jpg *.jpeg *.bmp)"
        )
        return file_paths

    def update_batch_summary_label(self):
        self.batch_summary_label.setText(f"등록 이미지: {self.batch_table.rowCount()}장")

    def _append_batch_row(self, file_path, row_count):
        row_index = self.batch_table.rowCount()
        self.batch_table.insertRow(row_index)

        file_name_item = QTableWidgetItem(os.path.basename(file_path))
        file_name_item.setFlags(file_name_item.flags() & ~Qt.ItemIsEditable)
        self.batch_table.setItem(row_index, 0, file_name_item)

        path_item = QTableWidgetItem(file_path)
        path_item.setFlags(path_item.flags() & ~Qt.ItemIsEditable)
        self.batch_table.setItem(row_index, 1, path_item)

        row_count_item = QTableWidgetItem(str(row_count))
        self.batch_table.setItem(row_index, 2, row_count_item)

    def add_batch_images(self):
        file_paths = self.select_images_common("원본 이미지 여러 장 선택")
        if not file_paths:
            return

        existing_paths = {
            self.batch_table.item(row, 1).text()
            for row in range(self.batch_table.rowCount())
            if self.batch_table.item(row, 1)
        }

        added_count = 0
        default_row_count = self.row_spin.value()

        for file_path in file_paths:
            if file_path in existing_paths:
                continue
            self._append_batch_row(file_path, default_row_count)
            existing_paths.add(file_path)
            added_count += 1

        self.update_batch_summary_label()

        if added_count == 0:
            QMessageBox.information(self, "안내", "추가된 새 이미지가 없습니다.")
        else:
            self.log(f"이미지 {added_count}장을 목록에 추가했습니다.")

    def remove_selected_batch_images(self):
        selected_rows = sorted(
            {item.row() for item in self.batch_table.selectedItems()},
            reverse=True,
        )

        if not selected_rows:
            QMessageBox.information(self, "안내", "삭제할 이미지를 선택하세요.")
            return

        for row in selected_rows:
            self.batch_table.removeRow(row)

        self.update_batch_summary_label()
        self.log(f"선택 이미지 {len(selected_rows)}장을 목록에서 삭제했습니다.")

    def clear_batch_images(self):
        if self.batch_table.rowCount() == 0:
            return

        self.batch_table.setRowCount(0)
        self.update_batch_summary_label()
        self.log("이미지 목록을 비웠습니다.")

    def collect_batch_items(self):
        items = []

        if self.batch_table.rowCount() == 0:
            raise RuntimeError("원본 이미지를 한 장 이상 추가하세요.")

        for row in range(self.batch_table.rowCount()):
            path_item = self.batch_table.item(row, 1)
            row_count_item = self.batch_table.item(row, 2)

            image_path = path_item.text().strip() if path_item else ""
            row_text = row_count_item.text().strip() if row_count_item else ""

            if not image_path:
                raise RuntimeError(f"{row + 1}번째 이미지 경로가 비어 있습니다.")
            if not os.path.exists(image_path):
                raise RuntimeError(f"파일을 찾을 수 없습니다: {image_path}")

            try:
                row_count = int(row_text)
            except Exception:
                raise RuntimeError(f"{row + 1}번째 row 수가 숫자가 아닙니다: {row_text}")

            if row_count <= 0:
                raise RuntimeError(f"{row + 1}번째 row 수는 1 이상이어야 합니다.")

            items.append({
                "path": image_path,
                "row_count": row_count,
            })

        return items

    def fetch_guild_members(self):
        world_name = self.world_combo.currentText().strip()
        guild_name = self.guild_input.text().strip()

        if not guild_name:
            QMessageBox.warning(self, "안내", "길드 이름을 입력하세요.")
            return

        if self.guild_worker is not None and self.guild_worker.isRunning():
            QMessageBox.warning(self, "안내", "이미 길드원 불러오는 중입니다.")
            return

        self.official_names = []
        self.official_list.clear()

        self.set_ui_enabled(False)
        self.log_box.clear()
        self.log("공식 길드원 명단 불러오기 시작...")

        self.guild_worker = GuildFetchWorker(world_name, guild_name)
        self.guild_worker.log_signal.connect(self.log)
        self.guild_worker.done_signal.connect(self.on_guild_fetch_done)
        self.guild_worker.error_signal.connect(self.on_guild_fetch_error)
        self.guild_worker.start()

    def on_guild_fetch_done(self, names):
        self.official_names = names
        self.official_list.clear()

        for name in names:
            self.official_list.addItem(name)

        self.log(f"공식 길드원 불러오기 완료: 총 {len(names)}명")

        if self.accumulated_rows:
            self.apply_auto_correction()

        self.set_ui_enabled(True)
        QMessageBox.information(self, "완료", f"공식 길드원 {len(names)}명을 불러왔습니다.")

    def on_guild_fetch_error(self, message):
        self.set_ui_enabled(True)
        self.log(f"오류: {message}")
        QMessageBox.critical(self, "오류", message)

    def run_ocr(self):
        if self.ocr_worker is not None and self.ocr_worker.isRunning():
            QMessageBox.warning(self, "안내", "이미 OCR 처리 중입니다.")
            return

        try:
            batch_items = self.collect_batch_items()
        except Exception as e:
            QMessageBox.warning(self, "안내", str(e))
            return

        self.set_ui_enabled(False)
        self.log("목록 전체 OCR 처리를 시작합니다...")

        self.ocr_worker = BatchOCRWorker(batch_items)
        self.ocr_worker.log_signal.connect(self.log)
        self.ocr_worker.done_signal.connect(self.on_ocr_done)
        self.ocr_worker.error_signal.connect(self.on_ocr_error)
        self.ocr_worker.start()

    def on_ocr_done(self, rows, batch_summary):
        self.accumulated_rows.extend(rows)
        self.processed_batches.extend(batch_summary)
        self.populate_table(self.accumulated_rows)

        success_count = sum(1 for item in batch_summary if item["status"] == "완료")
        fail_count = len(batch_summary) - success_count

        self.log(f"OCR 완료: 이번 실행 {len(rows)}개 row")
        self.log(f"현재 누적 row 수: {len(self.accumulated_rows)}")
        self.log(f"처리 이미지 결과: 성공 {success_count}장 / 실패 {fail_count}장")

        if self.official_names:
            self.apply_auto_correction()

        self.set_ui_enabled(True)
        QMessageBox.information(
            self,
            "완료",
            (
                "목록 전체 OCR 처리가 끝났습니다.\n"
                f"성공 {success_count}장 / 실패 {fail_count}장\n"
                f"이번 실행 {len(rows)}개 / 누적 {len(self.accumulated_rows)}개"
            ),
        )

    def on_ocr_error(self, message):
        self.set_ui_enabled(True)
        self.log(f"오류: {message}")
        QMessageBox.critical(self, "오류", message)

    def populate_table(self, rows):
        self.table.setRowCount(len(rows))

        for i, row in enumerate(rows):
            self.table.setItem(i, 0, QTableWidgetItem(str(row.get("row", i + 1))))
            self.table.setItem(i, 1, QTableWidgetItem(row.get("ocr_nickname", "")))
            self.table.setItem(i, 2, QTableWidgetItem(row.get("corrected_nickname", "")))
            self.table.setItem(i, 3, QTableWidgetItem(row.get("match_score", "")))
            self.table.setItem(i, 4, QTableWidgetItem(row.get("suro", "")))
            self.table.setItem(i, 5, QTableWidgetItem(row.get("flag", "")))

    def apply_auto_correction(self):
        if not self.official_names:
            QMessageBox.warning(self, "안내", "먼저 공식 길드원 명단을 불러오세요.")
            return

        if not self.accumulated_rows:
            QMessageBox.warning(self, "안내", "먼저 OCR을 실행하세요.")
            return

        before = [row.get("corrected_nickname", row.get("ocr_nickname", "")) for row in self.accumulated_rows]

        self.accumulated_rows = run_global_unique_matching(
            self.accumulated_rows,
            self.official_names,
        )

        after = [row.get("corrected_nickname", "") for row in self.accumulated_rows]
        changed_count = sum(1 for a, b in zip(before, after) if a != b)

        self.populate_table(self.accumulated_rows)
        self.log(f"전역 1:1 자동 보정 완료: {changed_count}개 row 변경")

    def save_excel(self):
        if not self.accumulated_rows:
            QMessageBox.warning(self, "안내", "저장할 결과가 없습니다.")
            return

        world_name = self.world_combo.currentText().strip()
        guild_name = self.guild_input.text().strip() or "guild"

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "엑셀 저장",
            f"{guild_name}_ocr_result.xlsx",
            "Excel Files (*.xlsx)"
        )

        if not file_path:
            return

        if not file_path.endswith(".xlsx"):
            file_path += ".xlsx"

        try:
            rows_to_save = [dict(row) for row in self.accumulated_rows]

            if self.official_names:
                rows_to_save = run_global_unique_matching(rows_to_save, self.official_names)

            wb = Workbook()

            ws = wb.active
            ws.title = "결과"

            headers = ["매칭 점수", "OCR 닉네임", "보정 닉네임", "닉네임 이미지", "실제 닉네임", "수로 점수", "플래그 점수"]
            for col, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col, value=header)

            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 20
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 28
            ws.column_dimensions["E"].width = 20
            ws.column_dimensions["F"].width = 14
            ws.column_dimensions["G"].width = 14

            for row_idx, row_data in enumerate(rows_to_save, start=2):
                ws.cell(row=row_idx, column=1, value=row_data.get("match_score", ""))
                ws.cell(row=row_idx, column=2, value=row_data.get("ocr_nickname", ""))
                ws.cell(row=row_idx, column=3, value=row_data.get("corrected_nickname", ""))
                ws.cell(row=row_idx, column=5, value="")
                ws.cell(row=row_idx, column=6, value=excel_numeric_value(row_data.get("suro", "")))
                ws.cell(row=row_idx, column=7, value=excel_numeric_value(row_data.get("flag", "")))

                img_bytes = row_data.get("nickname_image_bytes", b"")
                if img_bytes:
                    try:
                        bio = BytesIO(img_bytes)
                        xl_img = XLImage(bio)

                        max_w = 180
                        max_h = 42

                        orig_w = xl_img.width
                        orig_h = xl_img.height

                        if orig_w > 0 and orig_h > 0:
                            scale = min(max_w / orig_w, max_h / orig_h)
                            scale = min(scale, 1.0)
                            xl_img.width = int(orig_w * scale)
                            xl_img.height = int(orig_h * scale)

                        ws.add_image(xl_img, f"D{row_idx}")

                        row_height_points = max(35, int(xl_img.height * 0.75))
                        ws.row_dimensions[row_idx].height = row_height_points

                    except Exception:
                        pass

                score_text = str(row_data.get("match_score", "")).strip()

                try:
                    score_value = float(score_text) if score_text else 0.0
                except Exception:
                    score_value = 0.0

                actual_name_cell = ws.cell(row=row_idx, column=5)

                if score_text:
                    if score_value >= 0.75:
                        actual_name_cell.fill = LIGHT_PURPLE_FILL
                    elif score_value < 0.35:
                        actual_name_cell.fill = LIGHT_GRAY_FILL

            info = wb.create_sheet("입력 정보")
            info["A1"] = "항목"
            info["B1"] = "내용"
            info["A2"] = "월드"
            info["B2"] = world_name
            info["A3"] = "길드 이름"
            info["B3"] = guild_name
            info["A4"] = "공식 길드원 수"
            info["B4"] = len(self.official_names)
            info["A5"] = "누적 OCR row 수"
            info["B5"] = len(rows_to_save)
            info["A6"] = "등록된 이미지 수"
            info["B6"] = self.batch_table.rowCount()
            info["A7"] = "성공 처리 이미지 수"
            info["B7"] = sum(1 for item in self.processed_batches if item.get("status") == "완료")

            if self.processed_batches:
                batch_ws = wb.create_sheet("처리 이미지 목록")
                batch_headers = ["파일명", "전체 경로", "입력 row 수", "생성 row 수", "상태"]
                for col, header in enumerate(batch_headers, start=1):
                    batch_ws.cell(row=1, column=col, value=header)

                batch_ws.column_dimensions["A"].width = 28
                batch_ws.column_dimensions["B"].width = 60
                batch_ws.column_dimensions["C"].width = 12
                batch_ws.column_dimensions["D"].width = 12
                batch_ws.column_dimensions["E"].width = 28

                for idx, item in enumerate(self.processed_batches, start=2):
                    batch_ws.cell(row=idx, column=1, value=item.get("image_name", ""))
                    batch_ws.cell(row=idx, column=2, value=item.get("image_path", ""))
                    batch_ws.cell(row=idx, column=3, value=item.get("row_count", ""))
                    batch_ws.cell(row=idx, column=4, value=item.get("produced_rows", ""))
                    batch_ws.cell(row=idx, column=5, value=item.get("status", ""))

            if self.official_names:
                ws2 = wb.create_sheet("공식 길드원 명단")
                ws2["A1"] = "번호"
                ws2["B1"] = "길드원 이름"
                for i, name in enumerate(self.official_names, start=2):
                    ws2[f"A{i}"] = i - 1
                    ws2[f"B{i}"] = name

            wb.save(file_path)
            QMessageBox.information(self, "완료", f"엑셀 저장 완료\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "오류", str(e))

    def reset_all(self):
        if self.guild_worker is not None and self.guild_worker.isRunning():
            QMessageBox.warning(self, "안내", "길드원 불러오는 중에는 초기화할 수 없습니다.")
            return
        if self.ocr_worker is not None and self.ocr_worker.isRunning():
            QMessageBox.warning(self, "안내", "OCR 처리 중에는 초기화할 수 없습니다.")
            return

        self.official_names = []
        self.accumulated_rows = []
        self.processed_batches = []

        self.guild_input.clear()
        self.official_list.clear()
        self.log_box.clear()
        self.table.clearContents()
        self.table.setRowCount(0)
        self.batch_table.setRowCount(0)
        self.update_batch_summary_label()

        QMessageBox.information(self, "초기화", "초기화가 완료되었습니다.")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = IntegratedApp()
    window.show()
    sys.exit(app.exec())
